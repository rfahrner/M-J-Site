#!/usr/bin/env python3
"""
Import historic Atlanta / Building C / Delaware workbooks (and the
Houston CSV) into Supabase.

USAGE:
    pip install pandas openpyxl supabase --break-system-packages
    export SUPABASE_URL="https://ygsapysqzwrpcimgvaqx.supabase.co"
    export SUPABASE_KEY="sb_publishable_..."   # anon/publishable is fine, RLS allows insert

    python3 import_historics.py --dry-run Copy_of_Atlanta_workbook_FULL.xlsx
    python3 import_historics.py Copy_of_Atlanta_workbook_FULL.xlsx
    python3 import_historics.py --location buildingc Building_C_copy_FULL.xlsx
    python3 import_historics.py --location delaware Delaware_FULL.xlsx
    python3 import_historics.py --houston --date 2026-07-04 Kroger_Houston_....csv

ALWAYS run with --dry-run first. It prints exactly what would be
imported (tab-by-tab row counts, and any tabs it can't confidently
parse) without writing anything to Supabase.
"""

import argparse
import os
import re
import sys
from datetime import date

import pandas as pd
from openpyxl import load_workbook

try:
    from supabase import create_client
except ImportError:
    create_client = None  # only required for a real (non --dry-run) run

SKIP_TAB_NAMES = {"copy only", "schedule", "schedule del", "available"}
HEADER_ROW_INDEX = 4  # 0-indexed — row 5 in Excel's own numbering

# Tabs whose names don't fit the standard pattern, confirmed by hand.
# Add more here as they come up — checked before the regex parser runs.
MANUAL_TAB_OVERRIDES = {
    # name: (location, date, structure)
    # structure = which column layout to read this tab with. Normally same
    # as location, but 06-28 (Bill Schneider) has Delaware-style columns
    # (Customer rate/Carrier rate at 8/9) even though the shift itself
    # belongs under Atlanta — confirmed from the header-mismatch dry run.
    "06-28 (Bill Schneider)": ("atlanta", "2026-06-28", "delaware"),
    "5-31 C & Hostler": ("buildingc", "2026-05-31", "buildingc"),
    "Fathers Day (Sun)": ("atlanta", "2026-06-21", "atlanta"),
}

# These tabs span multiple days in one sheet — same column layout as every
# other tab, but each ROW carries its own date (column 6) instead of the
# whole tab representing a single day. Confirmed against the actual sheet
# rather than assumed.
PER_ROW_DATE_TABS = {"MTY to Kenlake Foods", "Dairy Emergency", "ATL to Murray, KY"}

def parse_row_date(row, idx=6):
    """Reads the per-row Date cell (col 6) for PER_ROW_DATE_TABS. Returns
    'YYYY-MM-DD' or None if that row has no date filled in."""
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if hasattr(v, "date") and callable(getattr(v, "date")):
        return v.date().isoformat()          # datetime.datetime
    if hasattr(v, "isoformat"):
        return v.isoformat()                  # already a datetime.date
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', str(v).strip())
    if m:
        try:
            return date(int(m.group(3)), int(m.group(1)), int(m.group(2))).isoformat()
        except ValueError:
            return None
    return None

# ---------------- tab name -> (location, date) ----------------

def parse_tab_name(name, default_location="atlanta"):
    """Returns (location_key, 'YYYY-MM-DD', structure_key) or None if unrecognized."""
    n = name.strip()
    if n in MANUAL_TAB_OVERRIDES:
        return MANUAL_TAB_OVERRIDES[n]
    if n.lower() in SKIP_TAB_NAMES:
        return None

    location = default_location
    date_part = n
    m = re.match(r'^(.*?)\s*(C|DEL)$', n, re.IGNORECASE)
    if m and m.group(1).strip():
        date_part = m.group(1).strip()
        location = "buildingc" if m.group(2).upper() == "C" else "delaware"

    # MM-DD (assume year 2026, per the sheet owner's confirmation)
    m = re.match(r'^(\d{1,2})-(\d{1,2})$', date_part)
    if m:
        try:
            return location, date(2026, int(m.group(1)), int(m.group(2))).isoformat(), location
        except ValueError:
            return None

    # M.DD.YY
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{2})$', date_part)
    if m:
        try:
            return location, date(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2))).isoformat(), location
        except ValueError:
            return None

    return None  # doesn't match a known pattern — flagged for manual review, not guessed at


# ---------------- header discovery for one tab ----------------

def norm(h):
    if h is None:
        return ""
    return re.sub(r'\s+', ' ', str(h)).strip()

# Column index map verified against the actual uploaded sample tabs
# (Atlanta 07-09, Building C 06-29 C, Delaware 7-1 DEL). Validated
# against each tab's own header row before use — see validate_headers().
SHIFT_COLS = {
    "interchange_ins_snapshot": 0, "interchange_agreement_snapshot": 1,
    "email_snapshot": 3, "dispatcher_phone_snapshot": 4, "aljex_load_number": 5,
    "sheet_date_cell": 6, "mc_snapshot": 7,
    "col8": 8, "col9": 9,  # meaning of 8/9 depends on location — see below
    "driver_name": 10, "driver_cell_snapshot": 11, "shift_start": 12,
    "pre_shift_text_sent": 13, "eta_for_shift_report": 14, "actual_shift_report": 15,
    "eta_for_next_dispatch": 16, "hos_time_left": 17, "comments": 18,
    "eta_disp_decimal": 19, "waynes_window": 20, "current_route_status": 21,
    "current_route_backhaul_status": 22, "trip1_call_time": 23,
    "hos_decimal": 24, "rev_level": 25, "next_call_time": 26,
}
TRIP1_COLS = {
    "route_id": 27, "trip_id": 28, "trailer_out": 29, "route_miles": 30,
    "stop_count": 31, "dispatch_time": 32, "last_stop_depart": 33, "return_to_dc": 34,
    "salvage": 35, "backhaul": 36, "backhaul_location": 38,
    "salvage_bhaul_refused_by": 39, "backhaul_trailer_number": 40, "return_eta_to_dc": 41,
    "return_drop_location": 42, "ppwk_received": 43, "time_sheet_start": 44,
    "time_sheet_end": 45, "drop_location_text": 46, "return_to_dc_text": 47,
    "backhaul_type": 48, "check_in_45min": 49, "text_not_answered_remind": 50,
    "pre_shift_call": 51, "route_est_hours": 52, "time_to_final_stop": 53,
    "eta_to_final_stop": 54, "est_route_complete": 55,
}
# Trips 2-5 share a reduced field set (no time-sheet/pre-shift-call tracking).
TRIP_N_TEMPLATE = {
    "route_id": 0, "trip_id": 1, "trailer_out": 2, "route_miles": 3, "stop_count": 4,
    "dispatch_time": 5, "last_stop_depart": 6, "return_to_dc": 7, "salvage": 8,
    "backhaul": 9, "backhaul_location": 11, "salvage_bhaul_refused_by": 12,
    "backhaul_trailer_number": 13, "return_eta_to_dc": 14, "return_drop_location": 15,
    "ppwk_received": 16, "route_est_hours": 18, "time_to_final_stop": 19,
    "eta_to_final_stop": 20, "est_route_complete": 21,
}
TRIP_N_BASE = {2: 57, 3: 80, 4: 103, 5: 126}
TRIP_N_CALL_TIME = {2: 56, 3: 79, 4: 102, 5: 125}


def validate_headers(header_row, structure):
    """Anchor checks — if these don't hold, don't trust positional
    extraction for this tab. Returns list of problems (empty = OK)."""
    h = [norm(x) for x in header_row]
    problems = []
    def check(idx, expect_substr, label):
        if idx >= len(h) or expect_substr.lower() not in h[idx].lower():
            problems.append(f"col{idx} ({label}): expected to contain {expect_substr!r}, got {h[idx] if idx < len(h) else '<missing>'!r}")
    check(10, "Driver", "driver name column")
    check(27, "Route", "trip 1 Route ID")
    check(29, "Trailer", "trip 1 Trailer Out")
    check(57, "Route", "trip 2 Route ID")
    if structure == "delaware":
        check(8, "rate", "Delaware customer rate")
        check(9, "rate", "Delaware carrier rate")
    else:
        check(8, "Rating", "driver rating")
    return problems


# ---------------- per-row extraction ----------------

def cellval(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if pd.isna(v):
        return None
    if hasattr(v, "isoformat"):  # date/time/datetime objects from openpyxl
        return v.isoformat()
    return v


def extract_shift(row, location, shift_date, structure=None):
    if structure is None:
        structure = location
    driver_name = cellval(row, SHIFT_COLS["driver_name"])
    if not driver_name or not str(driver_name).strip():
        return None  # blank row — nothing to import
    rec = {
        "location": location,
        "shift_date": shift_date,
        "driver_name_text": str(driver_name).strip(),
        "shift_start": cellval(row, SHIFT_COLS["shift_start"]),
        "aljex_load_number": cellval(row, SHIFT_COLS["aljex_load_number"]),
        "driver_cell_snapshot": cellval(row, SHIFT_COLS["driver_cell_snapshot"]),
        "email_snapshot": cellval(row, SHIFT_COLS["email_snapshot"]),
        "dispatcher_phone_snapshot": cellval(row, SHIFT_COLS["dispatcher_phone_snapshot"]),
        "mc_snapshot": cellval(row, SHIFT_COLS["mc_snapshot"]),
        "interchange_ins_snapshot": cellval(row, SHIFT_COLS["interchange_ins_snapshot"]),
        "interchange_agreement_snapshot": cellval(row, SHIFT_COLS["interchange_agreement_snapshot"]),
        "pre_shift_text_sent": cellval(row, SHIFT_COLS["pre_shift_text_sent"]),
        "eta_for_shift_report": cellval(row, SHIFT_COLS["eta_for_shift_report"]),
        "actual_shift_report": cellval(row, SHIFT_COLS["actual_shift_report"]),
        "eta_for_next_dispatch": cellval(row, SHIFT_COLS["eta_for_next_dispatch"]),
        "hos_time_left": cellval(row, SHIFT_COLS["hos_time_left"]),
        "comments": cellval(row, SHIFT_COLS["comments"]),
        "eta_disp_decimal": cellval(row, SHIFT_COLS["eta_disp_decimal"]),
        "waynes_window": cellval(row, SHIFT_COLS["waynes_window"]),
        "current_route_status": cellval(row, SHIFT_COLS["current_route_status"]),
        "current_route_backhaul_status": cellval(row, SHIFT_COLS["current_route_backhaul_status"]),
        "hos_decimal": cellval(row, SHIFT_COLS["hos_decimal"]),
        "rev_level": cellval(row, SHIFT_COLS["rev_level"]),
        "next_call_time": cellval(row, SHIFT_COLS["next_call_time"]),
    }
    if structure == "delaware":
        rec["customer_rate"] = cellval(row, SHIFT_COLS["col8"])
        rec["carrier_rate"] = cellval(row, SHIFT_COLS["col9"])
    else:
        rec["driver_rating_snapshot"] = cellval(row, SHIFT_COLS["col8"])
        rec["long_haul_flag"] = cellval(row, SHIFT_COLS["col9"])
    return rec


def extract_trips(row):
    """Returns list of (trip_number, dict) for whichever of the 5 slots
    actually have a Route ID filled in — blank slots are skipped, not
    inserted as empty trip rows."""
    trips = []
    if cellval(row, TRIP1_COLS["route_id"]) is not None:
        rec = {name: cellval(row, idx) for name, idx in TRIP1_COLS.items()}
        rec["call_time"] = cellval(row, SHIFT_COLS["trip1_call_time"])
        trips.append((1, rec))
    for n in (2, 3, 4, 5):
        base = TRIP_N_BASE[n]
        route_idx = base + TRIP_N_TEMPLATE["route_id"]
        if cellval(row, route_idx) is None:
            continue
        rec = {name: cellval(row, base + off) for name, off in TRIP_N_TEMPLATE.items()}
        rec["call_time"] = cellval(row, TRIP_N_CALL_TIME[n])
        trips.append((n, rec))
    return trips


# ---------------- driver lookup ----------------

def build_driver_lookup(supabase):
    resp = supabase.table("atlanta_drivers").select("id, \"Driver Name\"").execute()
    return {(r["Driver Name"] or "").strip().lower(): r["id"] for r in resp.data}


# ---------------- main workbook import ----------------

def import_workbook(path, default_location, dry_run, supabase):
    wb = load_workbook(path, read_only=True, data_only=True)
    driver_lookup = build_driver_lookup(supabase) if (supabase and not dry_run) else {}

    total_shifts = 0
    total_trips = 0
    skipped_tabs = []

    for sheet_name in wb.sheetnames:
        clean_name = sheet_name.strip()
        is_per_row_date = clean_name in PER_ROW_DATE_TABS

        if is_per_row_date:
            location = "atlanta"
            structure = "atlanta"
        else:
            parsed = parse_tab_name(sheet_name, default_location)
            if parsed is None:
                if clean_name.lower() not in SKIP_TAB_NAMES:
                    skipped_tabs.append(sheet_name)
                continue
            location, tab_shift_date, structure = parsed

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= HEADER_ROW_INDEX:
            skipped_tabs.append(f"{sheet_name} (too few rows)")
            continue

        problems = validate_headers(rows[HEADER_ROW_INDEX], structure)
        if problems:
            skipped_tabs.append(f"{sheet_name} (header mismatch: {problems[0]})")
            continue

        tab_shifts, tab_trips, undated_rows = 0, 0, 0
        for row in rows[HEADER_ROW_INDEX + 1:]:
            shift_date = parse_row_date(row) if is_per_row_date else tab_shift_date
            if shift_date is None:
                # only possible when is_per_row_date — a row with no date filled in yet
                undated_rows += 1
                continue

            shift = extract_shift(row, location, shift_date, structure)
            if shift is None:
                continue
            name_key = shift["driver_name_text"].strip().lower()
            shift["driver_id"] = driver_lookup.get(name_key)
            if shift["driver_id"] is not None:
                shift["driver_name_text"] = None  # have a real driver_id, don't also store free text

            trips = extract_trips(row)
            tab_shifts += 1
            tab_trips += len(trips)

            if not dry_run:
                inserted = supabase.table("loads_shifts").insert(shift).execute()
                shift_id = inserted.data[0]["id"]
                for trip_number, trip_rec in trips:
                    trip_rec["shift_id"] = shift_id
                    trip_rec["trip_number"] = trip_number
                    supabase.table("loads_trips").insert(trip_rec).execute()

        date_label = "(multiple dates)" if is_per_row_date else tab_shift_date
        undated_note = f", {undated_rows} row(s) had no date" if undated_rows else ""
        print(f"  {sheet_name:24s} -> {location:10s} {date_label:16s}   {tab_shifts:3d} shifts, {tab_trips:3d} loads{undated_note}")
        total_shifts += tab_shifts
        total_trips += tab_trips

    wb.close()
    print(f"\nTOTAL: {total_shifts} shifts, {total_trips} loads" + (" (DRY RUN — nothing written)" if dry_run else ""))
    if skipped_tabs:
        print(f"\n{len(skipped_tabs)} tab(s) skipped / need manual review:")
        for t in skipped_tabs:
            print(f"  - {t}")


# ---------------- Houston (flat CSV, no trip blocks) ----------------

def import_houston(path, shift_date, dry_run, supabase):
    df = pd.read_csv(path)
    df.columns = [norm(c) for c in df.columns]
    col_map = {
        "Aljex #": "aljex_number", "Comments": "comments", "TTC": "ttc", "TTT": "ttt",
        "Rating": "rating", "TIME": "time", "Name": "driver_name", "Phone #": "driver_phone",
        "TIME OUT | REMARKS": "time_out_remarks", "Dispatcher Phone #": "dispatcher_phone",
        "Carrier": "carrier", "MC": "mc", "Normal Rate": "normal_rate",
    }
    count = 0
    for _, row in df.iterrows():
        name = row.get("Name")
        if pd.isna(name) or not str(name).strip():
            continue
        rec = {"shift_date": shift_date}
        for src, dest in col_map.items():
            val = row.get(src)
            rec[dest] = None if pd.isna(val) else val
        count += 1
        if not dry_run:
            supabase.table("loads_houston").insert(rec).execute()
    print(f"Houston {shift_date}: {count} rows" + (" (DRY RUN)" if dry_run else ""))


# ---------------- CLI ----------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--location", default="atlanta", choices=["atlanta", "buildingc", "delaware"],
                     help="Default location for tabs with no C/DEL suffix (Atlanta tabs have none)")
    ap.add_argument("--houston", action="store_true", help="Treat file as a Houston CSV, not a workbook")
    ap.add_argument("--date", help="Required with --houston — the CSV has no date column")
    ap.add_argument("--dry-run", action="store_true", help="Preview only, writes nothing")
    args = ap.parse_args()

    supabase = None
    if not args.dry_run:
        if create_client is None:
            sys.exit("pip install supabase --break-system-packages (only needed for a real, non-dry-run import)")
        url = os.environ.get("SUPABASE_URL")
        key = os.environ.get("SUPABASE_KEY")
        if not url or not key:
            sys.exit("Set SUPABASE_URL and SUPABASE_KEY environment variables first.")
        supabase = create_client(url, key)

    if args.houston:
        if not args.date:
            sys.exit("--houston requires --date YYYY-MM-DD (the CSV filename's date is ambiguous)")
        import_houston(args.file, args.date, args.dry_run, supabase)
    else:
        import_workbook(args.file, args.location, args.dry_run, supabase)


if __name__ == "__main__":
    main()